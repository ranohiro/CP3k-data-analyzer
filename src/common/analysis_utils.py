from IPython.display import display, clear_output
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
import io
import ipywidgets as widgets
import itertools
import math
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import re

OUTPUT_ROOT = Path.cwd() / "data" / "export"
SHEET_PLOTS = "plots"
SHEET_SUMMARY = "summary"
SHEET_GROUP_STATS = "group_stats"
SHEET_OUTLIERS = "outliers"
SHEET_SAMPLE_METRICS = "sample_metrics"
SHEET_METRIC_DEFINITIONS = "metric_definitions"
OUT_SUFFIX = "_with_QCplots"
REGRESSION_METHODS_ALL = ["OLS", "Deming", "TheilSen", "PassingBablok"]
REGRESSION_LABELS = {
    "OLS": "OLS（最小二乗）",
    "Deming": "Deming（両軸誤差）",
    "TheilSen": "Theil-Sen（ロバスト）",
    "PassingBablok": "Passing-Bablok（ノンパラメトリック）"
}
ID_COL_CANDIDATES = ["検体ID", "SampleID", "Sample ID", "ID"]
GROUP_COL_CANDIDATES = ["種別", "Group", "Type", "分類"]
VALUE_PREFIXES = ("処方",)


def setup_japanese_font():
    candidates = [
        "IPAexGothic", "IPAPGothic",
        "Noto Sans CJK JP", "Noto Sans JP",
        "Yu Gothic", "YuGothic",
        "Meiryo", "MS Gothic",
        "Hiragino Sans", "TakaoGothic"
    ]

    available = {f.name for f in fm.fontManager.ttflist}
    chosen = None

    for c in candidates:
        if c in available:
            chosen = c
            break

    if chosen is None:
        for name in sorted(available):
            if any(k in name for k in ["IPA", "Noto", "Gothic", "Meiryo", "Hiragino", "ゴシック", "明朝"]):
                chosen = name
                break

    if chosen:
        plt.rcParams["font.family"] = chosen

    plt.rcParams["axes.unicode_minus"] = False

def pick_col(df, candidates, default=None):
    for c in candidates:
        if c in df.columns:
            return c
    return default

def normalize_group_col(df, group_col):
    if group_col is None:
        return None

    if group_col not in df.columns:
        return None

    s = df[group_col]

    if s.notna().sum() == 0:
        return None

    nunique = s.nunique(dropna=True)

    if nunique <= 1:
        return None

    if pd.api.types.is_numeric_dtype(s):
        n = s.notna().sum()
        if nunique > min(20, max(5, n // 3)):
            return None

    return group_col

def detect_value_cols(df, id_col, group_col, prefixes=("処方",)):
    cols = [c for c in df.columns if any(str(c).startswith(p) for p in prefixes)]

    if cols:
        return cols

    numeric_cols = df.select_dtypes(include="number").columns.tolist()

    exclude_cols = [id_col]
    if group_col:
        exclude_cols.append(group_col)

    return [c for c in numeric_cols if c not in exclude_cols]

def safe_name(s):
    s = str(s)
    return re.sub(r'[\\/:*?"<>|]', "_", s)

def _bytes_from_uploaded_content(content):
    if content is None:
        return None

    if isinstance(content, (bytes, bytearray)):
        return bytes(content)

    if hasattr(content, "tobytes"):
        return content.tobytes()

    return bytes(content)

def normalize_uploaded_item(uploader_value):
    v = uploader_value

    if not v:
        return None

    if isinstance(v, (list, tuple)):
        return v[0]

    if isinstance(v, dict):
        return list(v.values())[0]

    raise TypeError(f"FileUpload.value の型が想定外: {type(v)}")

def save_uploaded_xlsx(uploader_value, save_as="uploaded_input.xlsx"):
    item = normalize_uploaded_item(uploader_value)

    if item is None:
        return None

    if isinstance(item, dict):
        content = item.get("content", None)
    else:
        try:
            content = item["content"]
        except Exception:
            content = getattr(item, "content", None)

    b = _bytes_from_uploaded_content(content)

    if b is None:
        raise KeyError("アップロードから content を取得できませんでした。")

    path = Path(save_as)

    with open(path, "wb") as f:
        f.write(b)

    return path

def make_output_dirs(output_root=OUTPUT_ROOT, input_stem=None):
    """
    実行ごとに日時付きフォルダを作成する。

    例：
    LATEST
      └─ 20260703_133628_uploaded_input
          ├─ excel
          ├─ plots_png
          ├─ tables
          └─ logs
    """
    base_root = Path(output_root)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if input_stem:
        folder_name = f"{timestamp}_{safe_name(input_stem)}"
    else:
        folder_name = timestamp

    root = base_root / folder_name

    excel_dir = root / "excel"
    plot_dir = root / "plots_png"
    table_dir = root / "tables"
    log_dir = root / "logs"

    for d in [root, excel_dir, plot_dir, table_dir, log_dir]:
        d.mkdir(parents=True, exist_ok=True)

    return {
        "root": root,
        "excel": excel_dir,
        "plots": plot_dir,
        "tables": table_dir,
        "logs": log_dir,
        "run_label": folder_name
    }

def save_table_csv(df, path):
    if df is None:
        df = pd.DataFrame()

    df = clean_df_for_excel(df)

    df.to_csv(
        path,
        index=False,
        encoding="utf-8-sig"
    )

def apply_value_range_filter(
    df,
    xcol,
    ycol,
    use_range=False,
    lo=None,
    hi=None,
    mode="both"
):
    if not use_range:
        return df.copy()

    if lo is None or hi is None:
        return df.copy()

    if lo > hi:
        lo, hi = hi, lo

    x = pd.to_numeric(df[xcol], errors="coerce")
    y = pd.to_numeric(df[ycol], errors="coerce")

    mx = pd.Series(True, index=df.index)
    my = pd.Series(True, index=df.index)

    mx &= x >= lo
    mx &= x <= hi
    my &= y >= lo
    my &= y <= hi

    if mode == "both":
        m = mx & my
    elif mode == "either":
        m = mx | my
    elif mode == "x_only":
        m = mx
    elif mode == "y_only":
        m = my
    else:
        m = mx & my

    return df.loc[m].copy()

def pearson_r(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)

    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]

    if len(x) < 2:
        return np.nan

    return float(np.corrcoef(x, y)[0, 1])

def regression_fit(
    x,
    y,
    method="OLS",
    deming_lambda=1.0,
    theilsen_max_pairs=40000,
    seed=0
):
    x = np.asarray(x, float)
    y = np.asarray(y, float)

    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    n = len(x)

    if n < 2:
        return np.nan, np.nan

    if method == "OLS":
        a, b = np.polyfit(x, y, 1)
        return float(a), float(b)

    if method == "Deming":
        lx = float(deming_lambda) if deming_lambda else 1.0

        xbar = x.mean()
        ybar = y.mean()

        sxx = np.mean((x - xbar) ** 2)
        syy = np.mean((y - ybar) ** 2)
        sxy = np.mean((x - xbar) * (y - ybar))

        if sxy == 0:
            a = 0.0
        else:
            a = (
                syy - lx * sxx
                + math.sqrt((syy - lx * sxx) ** 2 + 4 * lx * sxy * sxy)
            ) / (2 * sxy)

        b = ybar - a * xbar
        return float(a), float(b)

    if method == "TheilSen":
        rng = np.random.default_rng(seed)
        idx = np.arange(n)
        total_pairs = n * (n - 1) // 2
        slopes = []

        if total_pairs <= theilsen_max_pairs:
            for i in range(n - 1):
                dx = x[i + 1:] - x[i]
                dy = y[i + 1:] - y[i]
                mask = dx != 0
                slopes.extend((dy[mask] / dx[mask]).tolist())
        else:
            for _ in range(theilsen_max_pairs):
                i, j = rng.choice(idx, 2, replace=False)
                dx = x[j] - x[i]

                if dx == 0:
                    continue

                slopes.append((y[j] - y[i]) / dx)

        a = float(np.median(slopes)) if slopes else 0.0
        b = float(np.median(y - a * x))

        return a, b

    a, b = np.polyfit(x, y, 1)
    return float(a), float(b)

def passing_bablok_fit_with_ci(
    x,
    y,
    alpha=0.05,
    max_pairs=200000,
    seed=0
):
    x = np.asarray(x, float)
    y = np.asarray(y, float)

    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    n = len(x)

    empty_info = {
        "slope_ci_low": np.nan,
        "slope_ci_high": np.nan,
        "intercept_ci_low": np.nan,
        "intercept_ci_high": np.nan,
        "pb_n_slopes": 0,
        "pb_ci_method": ""
    }

    if n < 3:
        empty_info["pb_ci_method"] = "insufficient n"
        return np.nan, np.nan, empty_info

    rng = np.random.default_rng(seed)
    total_pairs = n * (n - 1) // 2
    slopes = []

    if total_pairs <= max_pairs:
        for i in range(n - 1):
            dx = x[i + 1:] - x[i]
            dy = y[i + 1:] - y[i]

            mask = dx != 0
            s = dy[mask] / dx[mask]
            s = s[np.isfinite(s)]
            s = s[s != -1]

            slopes.extend(s.tolist())

        sampled = False
    else:
        idx = np.arange(n)

        for _ in range(max_pairs):
            i, j = rng.choice(idx, 2, replace=False)
            dx = x[j] - x[i]

            if dx == 0:
                continue

            s = (y[j] - y[i]) / dx

            if np.isfinite(s) and s != -1:
                slopes.append(float(s))

        sampled = True

    if len(slopes) == 0:
        empty_info["pb_ci_method"] = "no valid slopes"
        return np.nan, np.nan, empty_info

    slopes = np.sort(np.asarray(slopes, dtype=float))
    N = len(slopes)
    K = int(np.sum(slopes < -1))

    if N % 2 == 1:
        idx_med = (N - 1) // 2 + K
        idx_med = max(0, min(idx_med, N - 1))
        slope = float(slopes[idx_med])
    else:
        idx1 = N // 2 - 1 + K
        idx2 = N // 2 + K

        idx1 = max(0, min(idx1, N - 1))
        idx2 = max(0, min(idx2, N - 1))

        slope = float((slopes[idx1] + slopes[idx2]) / 2)

    intercept = float(np.median(y - slope * x))

    z = 1.959963984540054
    C = z * math.sqrt(n * (n - 1) * (2 * n + 5) / 18)

    m1 = int(math.floor((N - C) / 2 + K))
    m2 = int(math.ceil((N + C) / 2 + K))

    m1 = max(0, min(m1, N - 1))
    m2 = max(0, min(m2, N - 1))

    slope_low = float(slopes[min(m1, m2)])
    slope_high = float(slopes[max(m1, m2)])

    int_candidates = [
        np.median(y - slope_low * x),
        np.median(y - slope_high * x)
    ]

    intercept_low = float(np.nanmin(int_candidates))
    intercept_high = float(np.nanmax(int_candidates))

    info = {
        "slope_ci_low": slope_low,
        "slope_ci_high": slope_high,
        "intercept_ci_low": intercept_low,
        "intercept_ci_high": intercept_high,
        "pb_n_slopes": int(N),
        "pb_ci_method": "approx_rank_based_sampled" if sampled else "approx_rank_based_all_pairs"
    }

    return slope, intercept, info

def regression_fit_info(
    x,
    y,
    method="OLS",
    deming_lambda=1.0,
    theilsen_max_pairs=40000,
    seed=0
):
    info = {
        "slope_ci_low": np.nan,
        "slope_ci_high": np.nan,
        "intercept_ci_low": np.nan,
        "intercept_ci_high": np.nan,
        "pb_n_slopes": np.nan,
        "pb_ci_method": ""
    }

    if method == "PassingBablok":
        a, b, pb_info = passing_bablok_fit_with_ci(
            x,
            y,
            alpha=0.05,
            max_pairs=200000,
            seed=seed
        )
        info.update(pb_info)
        return a, b, info

    a, b = regression_fit(
        x,
        y,
        method=method,
        deming_lambda=deming_lambda,
        theilsen_max_pairs=theilsen_max_pairs,
        seed=seed
    )

    return a, b, info

def mad(arr):
    arr = np.asarray(arr, float)
    med = np.nanmedian(arr)
    return np.nanmedian(np.abs(arr - med))

def compute_residuals(x, y, a, b):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    return y - (a * x + b)

def classify_outlier_level(abs_z, thresh=3.5):
    if not np.isfinite(abs_z):
        return "not_evaluable"

    if abs_z >= thresh * (5.0/3.5):
        return "strong_candidate"

    if abs_z >= thresh:
        return "candidate"

    if abs_z >= thresh * (2.5/3.5):
        return "mild_candidate"

    return "none"

def compute_pair_sample_metrics(df, id_col, group_col, xcol, ycol, a, b, z_thresh=3.5):
    has_group = group_col and group_col in df.columns

    cols = [id_col, xcol, ycol] + ([group_col] if has_group else [])

    sub = df[cols].dropna(subset=[xcol, ycol]).copy()

    if sub.empty:
        return sub

    x = pd.to_numeric(sub[xcol], errors="coerce").astype(float)
    y = pd.to_numeric(sub[ycol], errors="coerce").astype(float)

    sub["X"] = xcol
    sub["Y"] = ycol

    sub["diff_y_minus_x"] = y - x
    sub["abs_diff_yx"] = np.abs(sub["diff_y_minus_x"])
    sub["mean_xy"] = (x + y) / 2.0

    sub["rel_diff_yx_pct"] = np.where(
        sub["mean_xy"] != 0,
        sub["abs_diff_yx"] / np.abs(sub["mean_xy"]) * 100,
        np.nan
    )

    sub["bias_pct_vs_x"] = np.where(
        x != 0,
        (y - x) / np.abs(x) * 100,
        np.nan
    )

    sub["ratio_y_over_x"] = np.where(
        x != 0,
        y / x,
        np.nan
    )

    sub["pred_y"] = a * x + b
    sub["residual"] = y - sub["pred_y"]
    sub["abs_residual"] = np.abs(sub["residual"])

    resid = sub["residual"].to_numpy(dtype=float)
    s = mad(resid)
    scale = 1.4826 * s if (np.isfinite(s) and s > 0) else np.nan

    if np.isfinite(scale):
        sub["z_MAD"] = sub["residual"] / scale
    else:
        sub["z_MAD"] = np.nan

    sub["abs_z_MAD"] = np.abs(sub["z_MAD"])
    sub["outlier_level"] = sub["abs_z_MAD"].apply(lambda z: classify_outlier_level(z, thresh=z_thresh))

    sub["outlier_basis"] = "回帰残差 y-(slope*x+intercept) をMADで標準化"
    sub["outlier_note"] = "乖離候補であり、自動除外ではない"

    return sub

def pick_outliers_table(
    df,
    id_col,
    group_col,
    xcol,
    ycol,
    a,
    b,
    z_thresh=3.5,
    top_n=200
):
    sub = compute_pair_sample_metrics(
        df=df,
        id_col=id_col,
        group_col=group_col,
        xcol=xcol,
        ycol=ycol,
        a=a,
        b=b
    )

    if sub.empty or "abs_z_MAD" not in sub.columns:
        return sub, sub

    flagged = sub[
        np.isfinite(sub["abs_z_MAD"])
        & (sub["abs_z_MAD"] >= z_thresh)
    ].copy()

    flagged = flagged.sort_values("abs_z_MAD", ascending=False).head(top_n)

    return sub, flagged

def plot_suite(
    df,
    id_col,
    group_col,
    xcol,
    ycol,
    method,
    lam,
    a,
    b,
    r,
    fit_info=None,
    z_thresh=3.5,
    show_outlier_labels=True,
    outlier_label_top=8,
    add_diag=True,
    show_fit=True,
    show_stats=True,
    show_equation=True,
    show_ci=True,
    show_ba_text=True,
    show_outlier_text=True,
    normal_s=16,
    outlier_s=28,
    outlier_lw=1.2,
    fig_width=16,
    fig_height=10,
    dpi=150,
    external_colors=None
):
    if fit_info is None:
        fit_info = {}
    if external_colors is not None: external_colors = np.asarray(external_colors)
    if external_colors is not None: external_colors = np.asarray(external_colors)


    has_group = group_col and group_col in df.columns

    base_cols = [id_col, xcol, ycol] + ([group_col] if has_group else [])

    sub = df[base_cols].dropna(subset=[xcol, ycol]).copy()

    if sub.empty:
        return None, None, None, None, None

    all_rows, flagged = pick_outliers_table(
        df=df,
        id_col=id_col,
        group_col=group_col,
        xcol=xcol,
        ycol=ycol,
        a=a,
        b=b,
        z_thresh=z_thresh,
        top_n=200
    )

    x = pd.to_numeric(sub[xcol], errors="coerce").astype(float).to_numpy()
    y = pd.to_numeric(sub[ycol], errors="coerce").astype(float).to_numpy()

    resid = compute_residuals(x, y, a, b)
    mean_xy = (x + y) / 2.0
    diff_yx = y - x

    bias = float(np.nanmean(diff_yx))
    sd = float(np.nanstd(diff_yx, ddof=1)) if len(diff_yx) > 1 else np.nan

    loa_hi = bias + 1.96 * sd if np.isfinite(sd) else np.nan
    loa_lo = bias - 1.96 * sd if np.isfinite(sd) else np.nan

    flagged_ids = set(flagged[id_col].tolist()) if flagged is not None and not flagged.empty else set()
    is_flagged = sub[id_col].isin(flagged_ids).to_numpy()

    if has_group:
        groups = sorted(sub[group_col].dropna().unique().tolist(), key=lambda t: str(t))
        cmap = plt.get_cmap("tab10")
        color_map = {g: cmap(i % 10) for i, g in enumerate(groups)}
        colors = sub[group_col].map(lambda g: color_map.get(g, "gray")).to_numpy()
    else:
        groups = []
        color_map = {}
        colors = np.array(["C0"] * len(sub))

    fig = plt.figure(figsize=(fig_width, fig_height), dpi=dpi)
    gs = fig.add_gridspec(2, 2)

    # ---------------- Scatter ----------------
    ax1 = fig.add_subplot(gs[0, 0])

    ax1.scatter(
        x[~is_flagged],
        y[~is_flagged],
        s=normal_s,
        alpha=0.75,
        c=external_colors[~is_flagged] if external_colors is not None else colors[~is_flagged]
    )

    if np.any(is_flagged):
        ax1.scatter(
            x[is_flagged],
            y[is_flagged],
            s=outlier_s,
            alpha=0.95,
            facecolors="none",
            edgecolors="red",
            linewidths=outlier_lw
        )

    ax1.set_title(f"散布図：{ycol} vs {xcol} / {REGRESSION_LABELS.get(method, method)}")
    ax1.set_xlabel(xcol)
    ax1.set_ylabel(ycol)
    ax1.grid(True, alpha=0.25)

    lo_xy = float(np.nanmin([np.nanmin(x), np.nanmin(y)]))
    hi_xy = float(np.nanmax([np.nanmax(x), np.nanmax(y)]))

    if add_diag:
        ax1.plot(
            [lo_xy, hi_xy],
            [lo_xy, hi_xy],
            "--",
            lw=1,
            alpha=0.6,
            label="y=x"
        )

    if show_fit and np.isfinite(a) and np.isfinite(b):
        xx = np.array([lo_xy, hi_xy])
        ax1.plot(
            xx,
            a * xx + b,
            lw=1.8,
            alpha=0.85,
            label="回帰直線"
        )

    if show_stats:
        lines = []
        lines.append(f"n={len(x)}")
        lines.append(f"Pearson r={r:.4f}")
        lines.append(
            REGRESSION_LABELS.get(method, method)
            + (f" (λ={lam:g})" if method == "Deming" else "")
        )

        if show_equation:
            lines.append(f"y={a:.4f}x+{b:.4f}")

        if show_ci:
            sl = fit_info.get("slope_ci_low", np.nan)
            sh = fit_info.get("slope_ci_high", np.nan)
            il = fit_info.get("intercept_ci_low", np.nan)
            ih = fit_info.get("intercept_ci_high", np.nan)

            if np.isfinite(sl) and np.isfinite(sh):
                lines.append(f"slope 95%CI [{sl:.4f}, {sh:.4f}]")

            if np.isfinite(il) and np.isfinite(ih):
                lines.append(f"intercept 95%CI [{il:.4f}, {ih:.4f}]")

        if show_ba_text:
            lines.append(f"BA bias={bias:.4g}")

            if np.isfinite(loa_lo) and np.isfinite(loa_hi):
                lines.append(f"LoA [{loa_lo:.4g}, {loa_hi:.4g}]")

        if show_outlier_text:
            n_flag = int(len(flagged)) if flagged is not None else 0
            lines.append(f"乖離候補: |z_MAD|≥{z_thresh}")
            lines.append(f"乖離候補数={n_flag}")

        ax1.text(
            0.03,
            0.97,
            "\n".join(lines),
            transform=ax1.transAxes,
            va="top",
            fontsize=9,
            bbox=dict(boxstyle="round", facecolor="white", alpha=0.75)
        )

    # ---------------- 乖離IDラベル ----------------
    if (
        show_outlier_labels
        and np.any(is_flagged)
        and outlier_label_top > 0
        and flagged is not None
        and not flagged.empty
    ):
        top_flagged = flagged.sort_values("abs_z_MAD", ascending=False).head(outlier_label_top)

        for _, row in top_flagged.iterrows():
            try:
                xx0 = float(row[xcol])
                yy0 = float(row[ycol])
                sid = row[id_col]

                if np.isfinite(xx0) and np.isfinite(yy0):
                    ax1.text(
                        xx0,
                        yy0,
                        str(sid),
                        fontsize=8,
                        color="red"
                    )
            except Exception:
                continue

    if has_group:
        for g in groups[:10]:
            ax1.scatter([], [], c=[color_map[g]], label=f"{group_col}:{g}")

    ax1.legend(fontsize=8, loc="best")

    # ---------------- Bland-Altman ----------------
    ax2 = fig.add_subplot(gs[0, 1])

    ax2.scatter(
        mean_xy[~is_flagged],
        diff_yx[~is_flagged],
        s=normal_s,
        alpha=0.75,
        c=external_colors[~is_flagged] if external_colors is not None else colors[~is_flagged]
    )

    if np.any(is_flagged):
        ax2.scatter(
            mean_xy[is_flagged],
            diff_yx[is_flagged],
            s=outlier_s,
            alpha=0.95,
            facecolors="none",
            edgecolors="red",
            linewidths=outlier_lw
        )

    ax2.axhline(
        bias,
        color="black",
        lw=1.5,
        label=f"平均差={bias:.3g}"
    )

    if np.isfinite(loa_hi) and np.isfinite(loa_lo):
        ax2.axhline(
            loa_hi,
            color="gray",
            lw=1.2,
            ls="--",
            label=f"+1.96SD={loa_hi:.3g}"
        )

        ax2.axhline(
            loa_lo,
            color="gray",
            lw=1.2,
            ls="--",
            label=f"-1.96SD={loa_lo:.3g}"
        )

    ax2.set_title("Bland–Altman：差(Y-X) vs 平均((X+Y)/2)")
    ax2.set_xlabel("平均 (X+Y)/2")
    ax2.set_ylabel("差 (Y-X)")
    ax2.grid(True, alpha=0.25)
    ax2.legend(fontsize=8, loc="best")

    # ---------------- Residuals ----------------
    ax3 = fig.add_subplot(gs[1, 0])

    ax3.scatter(
        x[~is_flagged],
        resid[~is_flagged],
        s=normal_s,
        alpha=0.75,
        c=external_colors[~is_flagged] if external_colors is not None else colors[~is_flagged]
    )

    if np.any(is_flagged):
        ax3.scatter(
            x[is_flagged],
            resid[is_flagged],
            s=outlier_s,
            alpha=0.95,
            facecolors="none",
            edgecolors="red",
            linewidths=outlier_lw
        )

    ax3.axhline(0, color="black", lw=1.2)
    ax3.set_title("残差プロット：残差(Y-(aX+b)) vs X")
    ax3.set_xlabel(xcol)
    ax3.set_ylabel("残差")
    ax3.grid(True, alpha=0.25)

    # ---------------- Text panel ----------------
    ax4 = fig.add_subplot(gs[1, 1])
    ax4.axis("off")

    n_flagged = int(len(flagged)) if flagged is not None else 0

    txt = (
        f"【統計まとめ】\n"
        f"X={xcol}\n"
        f"Y={ycol}\n"
        f"n={len(x)}\n"
        f"Pearson r={r:.6f}\n"
        f"回帰: {REGRESSION_LABELS.get(method, method)}"
        + (f" (λ={lam:g})" if method == "Deming" else "")
        + "\n"
        f"slope={a:.6f}\n"
        f"intercept={b:.6f}\n"
    )

    sl = fit_info.get("slope_ci_low", np.nan)
    sh = fit_info.get("slope_ci_high", np.nan)
    il = fit_info.get("intercept_ci_low", np.nan)
    ih = fit_info.get("intercept_ci_high", np.nan)

    if show_ci:
        if np.isfinite(sl) and np.isfinite(sh):
            txt += f"slope 95%CI=[{sl:.6g}, {sh:.6g}]\n"

        if np.isfinite(il) and np.isfinite(ih):
            txt += f"intercept 95%CI=[{il:.6g}, {ih:.6g}]\n"

    txt += (
        f"\n【Bland–Altman】\n"
        f"bias(Y-X)={bias:.6g}\n"
    )

    if np.isfinite(sd):
        txt += f"SD={sd:.6g}\nLoA=[{loa_lo:.6g}, {loa_hi:.6g}]\n"

    txt += (
        f"\n【乖離候補】\n"
        f"|z_MAD|≥{z_thresh}\n"
        f"乖離候補数={n_flagged}\n"
        f"※乖離候補は自動除外していません\n"
        f"※r/slope/interceptは解析対象全データで算出\n"
    )

    ax4.text(
        0.02,
        0.98,
        txt,
        va="top",
        fontsize=10
    )

    fig.tight_layout()

    return fig, sub, flagged, bias, (loa_lo, loa_hi)

def _excel_safe_value(v):
    if isinstance(v, (np.integer,)):
        return int(v)

    if isinstance(v, (np.floating,)):
        if np.isnan(v) or np.isinf(v):
            return None
        return float(v)

    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v

    if pd.isna(v):
        return None

    return v

def clean_df_for_excel(df):
    if df is None:
        return pd.DataFrame()

    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    out = out.where(pd.notna(out), None)

    return out

def insert_images_into_excel(
    input_xlsx,
    output_xlsx,
    image_paths,
    plot_sheet=SHEET_PLOTS,
    start_cell="A1",
    row_step=44,
    max_width_px=1100
):
    wb = load_workbook(input_xlsx)

    if plot_sheet in wb.sheetnames:
        wb.remove(wb[plot_sheet])

    ws = wb.create_sheet(plot_sheet)

    m = re.match(r"([A-Z]+)(\d+)", start_cell)

    if not m:
        start_cell = "A1"
        m = re.match(r"([A-Z]+)(\d+)", start_cell)

    col_letters = m.group(1)
    row = int(m.group(2))

    for p in image_paths:
        img = XLImage(str(p))

        if img.width and img.width > max_width_px:
            scale = max_width_px / img.width
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

        ws.add_image(img, f"{col_letters}{row}")
        row += row_step

    wb.save(output_xlsx)

def write_df_to_sheet(xlsx_path, df, sheet_name, index=False):
    df = clean_df_for_excel(df)

    wb = load_workbook(xlsx_path)

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    df_to_write = df.reset_index() if index else df

    ws.append(list(df_to_write.columns))

    for row in df_to_write.itertuples(index=False):
        ws.append([_excel_safe_value(v) for v in row])

    wb.save(xlsx_path)

def groupwise_stats(df, id_col, group_col, xcol, ycol, method, lam):
    if not (group_col and group_col in df.columns):
        return pd.DataFrame(columns=[
            "X",
            "Y",
            "group",
            "n",
            "r",
            "slope",
            "intercept",
            "regression",
            "lambda"
        ])

    rows = []

    sub = df[[group_col, xcol, ycol]].dropna(subset=[xcol, ycol]).copy()

    for g, gdf in sub.groupby(group_col, dropna=False):
        x = pd.to_numeric(gdf[xcol], errors="coerce").astype(float).to_numpy()
        y = pd.to_numeric(gdf[ycol], errors="coerce").astype(float).to_numpy()

        if len(x) < 2:
            continue

        r = pearson_r(x, y)
        a, b, fit_info = regression_fit_info(
            x,
            y,
            method=method,
            deming_lambda=lam
        )

        rows.append({
            "X": xcol,
            "Y": ycol,
            "group": str(g),
            "n": len(x),
            "r": r,
            "slope": a,
            "intercept": b,
            "regression": method,
            "regression_label": REGRESSION_LABELS.get(method, method),
            "lambda": lam if method == "Deming" else np.nan,
            "slope_95CI_low": fit_info.get("slope_ci_low", np.nan),
            "slope_95CI_high": fit_info.get("slope_ci_high", np.nan),
            "intercept_95CI_low": fit_info.get("intercept_ci_low", np.nan),
            "intercept_95CI_high": fit_info.get("intercept_ci_high", np.nan),
            "outlier_excluded_from_fit": False,
            "fit_data_note": "群別統計も乖離候補を除外せず算出"
        })

    return pd.DataFrame(rows)

def make_metric_definitions_df():
    rows = [
        ["入力データ", "検体ID / SampleID / ID", "各検体を識別する列", "入力列", "図中IDラベル、outliers、sample_metricsで検体を追跡", "列名がない場合は先頭列をID列として扱います"],
        ["入力データ", "種別 / Group / Type", "任意の分類列", "入力列", "存在し、分類列と判断できる場合のみ色分け・群別統計に使用", "なくても動きます"],
        ["入力データ", "処方列", "比較対象となる数値列", "入力列", "列名が「処方」で始まる列を優先。なければ数値列を候補", "2列以上必要です"],
        ["基本情報", "X", "比較ペアにおける基準側の列名", "列名", "例：X=処方1", ""],
        ["基本情報", "Y", "比較ペアにおける比較側の列名", "列名", "例：Y=処方2", ""],
        ["差分", "diff_y_minus_x", "YとXの単純差", "Y - X", "正ならYがXより高く、負なら低い", "Bland–Altmanの差と同じです"],
        ["差分", "abs_diff_yx", "YとXの差の絶対値", "|Y - X|", "方向を問わず離れ具合を見る", "単位依存です"],
        ["Bland–Altman", "mean_xy", "XとYの平均値", "(X + Y) / 2", "BAプロットの横軸", ""],
        ["相対差", "rel_diff_yx_pct", "平均値基準の相対差", "|Y - X| / |(X+Y)/2| × 100", "値域差をならしたズレ", "平均が0に近いと不安定"],
        ["相対差", "bias_pct_vs_x", "X基準の差の割合", "(Y - X) / |X| × 100", "Xに対してYが何%高い/低いか", "Xが0に近いと不安定"],
        ["比率", "ratio_y_over_x", "Y/X比", "Y / X", "1に近いほど近い", "Xが0に近いと不安定"],
        ["回帰", "pred_y", "回帰式から予測されるY", "slope × X + intercept", "全体傾向から期待されるY", ""],
        ["回帰残差", "residual", "実測Yと予測Yの差", "Y - pred_y", "回帰直線からのズレ", "単純なY-Xではありません"],
        ["回帰残差", "abs_residual", "残差の絶対値", "|Y - pred_y|", "方向を問わず回帰からの外れ具合", "単位依存です"],
        ["乖離候補", "z_MAD", "残差をMADで標準化した指標", "residual / (1.4826 × MAD(residual))", "全体の残差ばらつきに対する外れ具合", "自動除外基準ではありません"],
        ["乖離候補", "abs_z_MAD", "z_MADの絶対値", "|z_MAD|", "大きいほど乖離候補", "方向はz_MADまたはresidualで確認"],
        ["乖離候補", "outlier_level", "確認優先度", "2.5以上=mild, 3.5以上=candidate, 5.0以上=strong", "乖離候補の段階表示", "除外判定ではありません"],
        ["回帰統計", "r", "Pearson相関係数", "corr(X, Y)", "1に近いほど正の直線相関が強い", "一致性ではありません"],
        ["回帰統計", "slope", "回帰直線の傾き", "回帰法に依存", "1に近いほど比例性が近い", "回帰法により変わります"],
        ["回帰統計", "intercept", "回帰直線の切片", "回帰法に依存", "0に近いほど定数差が小さい", "測定範囲外の解釈注意"],
        ["回帰統計", "slope_95CI_low/high", "傾きの95%信頼区間", "Passing-Bablokで算出", "傾き推定の不確かさ", "現コードでは主にPassing-Bablokのみ"],
        ["回帰統計", "intercept_95CI_low/high", "切片の95%信頼区間", "Passing-Bablokで算出", "切片推定の不確かさ", "現コードでは主にPassing-Bablokのみ"],
        ["Bland–Altman", "BA_bias(y-x)", "Y-Xの平均差", "mean(Y - X)", "平均的な差", "外れ値の影響を受けます"],
        ["Bland–Altman", "BA_LoA_low / BA_LoA_high", "95%一致限界", "bias ± 1.96×SD(Y-X)", "差が多く入る想定範囲", "差の分布に依存"],
        ["仕様", "outlier_excluded_from_fit", "乖離候補を回帰計算から除外したか", "False=除外していない", "r/slope/interceptは解析対象全データで算出", "対象範囲フィルタで外れた値は除外"]
    ]

    return pd.DataFrame(
        rows,
        columns=["分類", "指標名", "意味", "計算式", "解釈", "注意点"]
    )
